import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter 

def ajustar_formato(valor):
    if isinstance(valor, str):
        valor = valor.replace('.', '').replace(',', '.')
    return valor      

def map_capacidade(row):
    if row['tipo_movel'] == 'ILHA':
        return row['ilha']
    elif row['tipo_movel'] == 'PG':
        return row['pg']
    elif row['tipo_movel'] == 'PE':
        return row['pe']
    else:
        return None

def analisar_desvio_quantidade(row):
    if row['qtd_expositor_total'] > row['venda_v30']:
        return '⚠️ Atenção Expositor Maior que a venda!'
    else:
        return '✅ Ok'

# Guerrilha - ITENS MAPEADOS

# guerrilha_df = pd.read_csv('guerrilha.csv')
guerrilha_df = pd.read_excel('linkker.xlsx')

# Exposições - Cadastro de exposições
expositor_df = pd.read_csv('exp.csv', sep=';',low_memory=False)

# Mix Total - Cadastro dos itens
mix_total_df = pd.read_csv('mix_total.csv', sep=';')


# Cadadatro de Usuário Abastecimento
# Informação do analista responsável no Abastecimento por departamento e seção
usuario_abastecimento = pd.read_excel('cadastro_usuario_abastecimento.xlsx')


lojas_sem_limitador = [51, 108, 116, 124, 167, 175, 183, 205, 230, 248, 264, 272, 310, 329, 337]  # Loja sem regra de corte (limitador da venda 15)

posicoes_compativeis = ['PE', 'PG', 'ILHA'] # Pontos que são reconhecidos como válidos

lista_de_produtos_em_linha = list(mix_total_df['RMS']) # lista dos itens em linha atualmente.

guerrilha_df['NÚMERO DA LOJA'] = guerrilha_df['NÚMERO DA LOJA'].astype(int)


colunas_manter_guerrilha = ['NÚMERO DA LOJA', 'TIPO DE PEX', 'NOME DO PONTO COMERCIALIZADO', 'SIGLA DO PONTO', 'Inicio', 'Fim',
                  'DATA DA VENDA', 'TIPO DE PERÍODO', 'USER QUE VENDEU', 'NOME SKU', 'CÓDIGO SKU']

guerrilha_df = guerrilha_df.loc[:, colunas_manter_guerrilha]

expositor_df['QTDVENDA_30_DIAS'] = pd.to_numeric(expositor_df['QTDVENDA_30_DIAS'].apply(ajustar_formato), errors='coerce')


colunas_manter_expositor = ['CODIGO', 'LOJA', 'ILHA_CAP', 'PE_CAP', 'PG_CAP', 'SALDO_CD', 'QTDVENDA_30_DIAS']

expositor_df = expositor_df.loc[:, colunas_manter_expositor]


colunas_renomear = {
    'CÓDIGO SKU': 'codigo_sku',
    'NÚMERO DA LOJA': 'loja',
    'TIPO DE PEX': 'nome_posicao',
    'NOME DO PONTO COMERCIALIZADO': 'posicao_movel',
    'SIGLA DO PONTO': 'tipo_movel',
    'Inicio': 'dt_inicio',
    'Fim': 'dt_fim',  
}

guerrilha_df.rename(columns=colunas_renomear, inplace= True)

guerrilha_df.drop(columns=['DATA DA VENDA','TIPO DE PERÍODO','USER QUE VENDEU','NOME SKU'], inplace=True)
guerrilha_df['loja'] = guerrilha_df['loja'].astype(int)


colunas_renomear = {
    'CODIGO': 'codigo',
    'LOJA': 'loja',
    'ILHA_CAP': 'ilha',
    'PE_CAP': 'pe',
    'PG_CAP': 'pg',
    'SALDO_CD': 'saldo_cd',
    'QTDVENDA_30_DIAS': 'venda_v30',
    
}

expositor_df.rename(columns=colunas_renomear, inplace= True)

expositor_df[['pe', 'pg', 'ilha', 'saldo_cd']] = expositor_df[['pe', 'pg', 'ilha', 'saldo_cd']].fillna(0)

#Juntando o mix_total_df com expositor_df

colunas_selecionadas_mix_total = ['DESCRICAO','SISTEMATICA', 'NOME_COMPRADOR', 'NOME_DEPARTAMENTO', 'SECAO', 'NOME_SECAO','TIPO_EMB_TRANSF','EAN']

expositor_mix_total = expositor_df.merge(mix_total_df[colunas_selecionadas_mix_total + ['RMS']],
                                        how = 'left',
                                        left_on='codigo',
                                        right_on='RMS')

expositor_mix_total.drop(columns='RMS', inplace=True)
expositor_mix_total.dropna(inplace=True)
expositor_mix_total[['SECAO','SISTEMATICA']] = expositor_mix_total[['SECAO','SISTEMATICA']].astype(int)

colunas_categoricas = ['NOME_SECAO', 'NOME_DEPARTAMENTO', 'NOME_COMPRADOR', 'DESCRICAO','TIPO_EMB_TRANSF','EAN']

for coluna in colunas_categoricas:
    expositor_mix_total[coluna] = expositor_mix_total[coluna].astype('category')

# Itens sem Exposição mapeados
    
itens_mapeados_exposicao = guerrilha_df.merge(expositor_mix_total,
                                              how ='left',
                                              left_on=['loja', 'codigo_sku'],
                                              right_on=['loja', 'codigo'],
                                             )

itens_mapeados_exposicao['capacidade_item_movel_mapeado'] = itens_mapeados_exposicao.apply(map_capacidade, axis=1)

itens_mapeados_exposicao = itens_mapeados_exposicao.dropna()

itens_mapeados_sem_cadastro_exposicao = itens_mapeados_exposicao[(itens_mapeados_exposicao['SISTEMATICA'] == 1) &
                                                                (itens_mapeados_exposicao['capacidade_item_movel_mapeado'] == 0 ) &
                                                                (itens_mapeados_exposicao['TIPO_EMB_TRANSF'] != 'KT') &
                                                                (~itens_mapeados_exposicao['EAN'].str.startswith('3'))
                                                                ]

itens_mapeados_sem_cadastro_exposicao = itens_mapeados_sem_cadastro_exposicao.merge(usuario_abastecimento,
                                                                                   how='left',
                                                                                   left_on= 'SECAO',
                                                                                   right_on= 'secao_id'
                                                                                   )

itens_mapeados_sem_cadastro_exposicao = itens_mapeados_sem_cadastro_exposicao.copy()
itens_mapeados_sem_cadastro_exposicao.dropna(inplace=True)
colunas_manter_relatorio_abastecimento = ['loja','tipo_movel','codigo_sku','DESCRICAO','capacidade_item_movel_mapeado','analista_abastecimento']


relatorio_itens_sem_cadastro_exposicao = itens_mapeados_sem_cadastro_exposicao.loc[:,colunas_manter_relatorio_abastecimento]

relatorio_itens_sem_cadastro_exposicao['Situação Item'] = '⚠️ Cadastrar Exposição!'

relatorio_itens_sem_cadastro_exposicao['capacidade_item_movel_mapeado'] = relatorio_itens_sem_cadastro_exposicao['capacidade_item_movel_mapeado'].astype(float)

nome_relatorio_abastecimento = 'itens_mapeados_sem_expositor_cadastrado.xlsx'

writer = pd.ExcelWriter(nome_relatorio_abastecimento, engine='xlsxwriter')



colunas_renomear_abastecimento ={
    'loja':'Loja', 
     'tipo_movel':'Móvel', 
     'codigo_sku':'Código Produto', 
     'DESCRICAO':'Descrição', 
     'capacidade_item_movel_mapeado':'Capacidade Cadastrado', 
     'analista_abastecimento':'Analista Abastecimento'
    }

relatorio_itens_sem_cadastro_exposicao.rename(columns=colunas_renomear_abastecimento, inplace =True )

analistas = relatorio_itens_sem_cadastro_exposicao['Analista Abastecimento'].unique()

for analista in analistas:
    analista_df = relatorio_itens_sem_cadastro_exposicao[relatorio_itens_sem_cadastro_exposicao['Analista Abastecimento'] == analista]
    analista_df.to_excel(writer, sheet_name=analista, index=False)

writer.close()

wb = load_workbook(f'{nome_relatorio_abastecimento}')


for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # celulas dos titulos das colunas
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="595959", end_color="595959", fill_type = "solid")
        cell.font = Font(color="D9D9D9")

    # bordas das celulas
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    alignment = Alignment(horizontal='center', vertical='center')
    for row in ws:
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment

    for i, col in enumerate(relatorio_itens_sem_cadastro_exposicao.columns):
        max_length = max(
            relatorio_itens_sem_cadastro_exposicao[col].astype(str).map(len).max(),
            len(str(col))
        )
        ws.column_dimensions[get_column_letter(i+1)].width = max_length

# Salve a planilha
wb.save('itens_mapeados_sem_expositor_cadastrado.xlsx')

print(f'{nome_relatorio_abastecimento} exportado com sucesso ✅')

# Relatório comercial

itens_mapeados_exposicao['capacidade_item_movel_mapeado'] = itens_mapeados_exposicao['capacidade_item_movel_mapeado'].apply(ajustar_formato).astype(float)

itens_mapeados_exposicao['saldo_cd'] = itens_mapeados_exposicao['saldo_cd'].apply(ajustar_formato).astype(float)
itens_mapeados_exposicao['tipo_movel'] = itens_mapeados_exposicao['tipo_movel'].astype('category')

itens_mapeados_exposicao.drop(columns='codigo', inplace= True)

relatorio_comercial = itens_mapeados_exposicao[
    (itens_mapeados_exposicao['saldo_cd'] > 0) &
    (itens_mapeados_exposicao['SISTEMATICA'] == 1) &
    (itens_mapeados_exposicao['tipo_movel'].isin(posicoes_compativeis)) &
    (itens_mapeados_exposicao['loja'].isin(lojas_sem_limitador)) &
    (itens_mapeados_exposicao['codigo_sku'].isin(lista_de_produtos_em_linha))
]

agupar_venda30_por_posicao = relatorio_comercial.groupby(['loja', 'posicao_movel'])['venda_v30'].sum().reset_index()
agupar_venda30_por_posicao = agupar_venda30_por_posicao.rename(columns={'venda_v30': 'venda_30_posicao'})

relatorio_comercial = pd.merge(relatorio_comercial,agupar_venda30_por_posicao, on=['loja', 'posicao_movel'], how='left')

relatorio_comercial['qtd_expositor_posicao'] = 0



for index, row in relatorio_comercial.iterrows():
    venda_30 = row['venda_v30']
    soma_venda_posicao = row['venda_30_posicao']
    capacidade_produto_movel_mapeado =  row['capacidade_item_movel_mapeado']
    saldo_cd = row['saldo_cd']

    if saldo_cd > 1:
        if soma_venda_posicao != 0:
                    relatorio_comercial.at[index,'percentual_rateio'] = (venda_30 / soma_venda_posicao)
                    qtd_exp_temp = (venda_30 / soma_venda_posicao) * capacidade_produto_movel_mapeado
                    relatorio_comercial.at[index, 'qtd_expositor_posicao'] = int(qtd_exp_temp)
relatorio_comercial['qtd_expositor_total'] = relatorio_comercial.groupby(['loja', 'codigo_sku'])['qtd_expositor_posicao'].transform('sum')



relatorio_comercial['qtd_mapeado_mesma_loja'] =  relatorio_comercial.groupby(['loja', 'codigo_sku'])['codigo_sku'].transform('count')


relatorio_comercial['qtd_expositor_posicao'] = 0



for index, row in relatorio_comercial.iterrows():
    venda_30 = row['venda_v30']
    soma_venda_posicao = row['venda_30_posicao']
    capacidade_produto_movel_mapeado =  row['capacidade_item_movel_mapeado']
    saldo_cd = row['saldo_cd']

    if saldo_cd > 1:
        if soma_venda_posicao != 0:
                    relatorio_comercial.at[index,'percentual_rateio'] = (venda_30 / soma_venda_posicao)
                    qtd_exp_temp = (venda_30 / soma_venda_posicao) * capacidade_produto_movel_mapeado
                    relatorio_comercial.at[index, 'qtd_expositor_posicao'] = int(qtd_exp_temp)
relatorio_comercial['qtd_expositor_total'] = relatorio_comercial.groupby(['loja', 'codigo_sku'])['qtd_expositor_posicao'].transform('sum')

relatorio_comercial['qtd_mapeado_mesma_loja'] =  relatorio_comercial.groupby(['loja', 'codigo_sku'])['codigo_sku'].transform('count')

relatorio_comercial['analisar'] = relatorio_comercial.apply(analisar_desvio_quantidade, axis=1)

relatorio_comercial.to_excel('relatório_agrupado_guerrilha.xlsx', index=False)

relatorio_comercial['percentual_rateio'] = relatorio_comercial['percentual_rateio'].round(2)

nome_relatorio_comercial = 'relatório_agrupado_guerrilha.xlsx'
colunas_nao_alinhar = ['DESCRICAO', 'analisar','NOME_SECAO', 'NOME_DEPARTAMENTO','nome_posicao']

col_to_letter = {col: get_column_letter(i+1) for i, col in enumerate(relatorio_comercial.columns)}
colunas_nao_alinhar = [col_to_letter[col] for col in colunas_nao_alinhar if col in col_to_letter]


wb = load_workbook(f'{nome_relatorio_comercial}')
ws = wb.active


for cell in ws[1]:
    cell.fill = PatternFill(start_color="595959", end_color="595959", fill_type = "solid")
    cell.font = Font(color="D9D9D9")


thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

alignment = Alignment(horizontal='center', vertical='center')
for row in ws:
    for cell in row:
        cell.border = thin_border
        if cell.column_letter not in colunas_nao_alinhar:
            cell.alignment = alignment    

for i, col in enumerate(relatorio_comercial.columns):
    max_length = max(
        relatorio_comercial[col].astype(str).map(len).max(),
        len(str(col))
    )
    ws.column_dimensions[get_column_letter(i+1)].width = max_length


wb.save(nome_relatorio_comercial)
wb.close();

print(f'{nome_relatorio_comercial} exportado com sucesso ✅')

# Relatório por comprador

colunas_manter_relatorio_por_comprador = ['NOME_COMPRADOR','loja','codigo_sku','DESCRICAO','nome_posicao','posicao_movel','tipo_movel','venda_v30','qtd_expositor_total','qtd_mapeado_mesma_loja','analisar']

relatorio_comercial_simplificado = relatorio_comercial[colunas_manter_relatorio_por_comprador]
relatorio_comercial_simplificado = relatorio_comercial_simplificado[relatorio_comercial_simplificado['venda_v30'] > 0]

relatorio_comercial_simplificado = relatorio_comercial_simplificado.copy()
relatorio_comercial_simplificado['quantidade_a_mais_venda_30'] = relatorio_comercial_simplificado['qtd_expositor_total'] / relatorio_comercial_simplificado['venda_v30']

relatorio_comercial_simplificado = relatorio_comercial_simplificado.sort_values(by='quantidade_a_mais_venda_30', ascending=False)


renomear_colunas_relatorio_comercial_simplificado = {'loja':'Loja',
                                                     'codigo_sku':'Código Produto',
                                                     'DESCRICAO':'Descrição',
                                                     'nome_posicao':'Nome da Posição',
                                                     'posicao_movel':'Posição Móvel',
                                                     'tipo_movel':'Tipo Móvel',
                                                     'venda_v30':'Venda 30 Dias',
                                                     'qtd_expositor_total':'Quantidade Expositor',
                                                     'qtd_mapeado_mesma_loja':'Quantidades de Pontos mapeados',
                                                     'analisar':'Analisar',
                                                     'NOME_COMPRADOR': 'Comprador'}

relatorio_comercial_simplificado.rename(columns=renomear_colunas_relatorio_comercial_simplificado,inplace=True)

relatorio_comercial_simplificado = relatorio_comercial_simplificado[relatorio_comercial_simplificado['quantidade_a_mais_venda_30'] > 1 ]

relatorio_comercial_simplificado = relatorio_comercial_simplificado.copy()
relatorio_comercial_simplificado.drop(columns=['quantidade_a_mais_venda_30'], inplace=True)

relatorio_comercial_simplificado.to_excel('relatorio_comprador.xlsx', index=False)

nome_relatorio_comercial_simplificado = 'relatorio_comprador.xlsx'
colunas_nao_alinhar_simplificado = ['Comprador', 'Descrição', 'nome_posicao']

grupos_comprador = relatorio_comercial_simplificado.groupby('Comprador',observed=False)

if not relatorio_comercial_simplificado.empty:
    writer = pd.ExcelWriter(nome_relatorio_comercial_simplificado, engine='openpyxl')

    for comprador, grupo in grupos_comprador:
        if not grupo.empty:
            nome_aba = comprador[:31]
            grupo.to_excel(writer, sheet_name=nome_aba, index=False)
            
    writer.close()

col_to_letter = {col: get_column_letter(i+1) for i, col in enumerate(relatorio_comercial_simplificado.columns)}
colunas_nao_alinhar_simplificado = [col_to_letter[col] for col in colunas_nao_alinhar_simplificado if col in col_to_letter]

wb = load_workbook(f'{nome_relatorio_comercial_simplificado}')

for ws in wb.worksheets:  # Itera sobre todas as planilhas no arquivo
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="595959", end_color="595959", fill_type = "solid")
        cell.font = Font(color="D9D9D9")

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    alignment = Alignment(horizontal='center', vertical='center')
    for row in ws:
        for cell in row:
            cell.border = thin_border
            if cell.column_letter not in colunas_nao_alinhar_simplificado:
                cell.alignment = alignment    

    for i, col in enumerate(relatorio_comercial_simplificado.columns):
        max_length = max(
            relatorio_comercial_simplificado[col].astype(str).map(len).max(),
            len(str(col))
        )
        ws.column_dimensions[get_column_letter(i+1)].width = max_length

wb.save(nome_relatorio_comercial_simplificado)
wb.close();

print(f'{nome_relatorio_comercial_simplificado} gerado com sucesso! ✅')
